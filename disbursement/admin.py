# -*- coding: utf-8 -*-
from __future__ import unicode_literals

from django.db.models import Q
from django.contrib import admin
from django.contrib.admin.templatetags.admin_urls import admin_urlname
from django.shortcuts import resolve_url
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _

from .mixins import (
    AdminSiteOwnerOnlyPermissionMixin,
    ExportCsvMixin,
    BankExportCsvMixin,
    BankExportExcelMixin,
)
from .models import (
    Agent,
    BankTransaction,
    DisbursementData,
    DisbursementDocData,
    VMTData,
    RemainingAmounts,
)
from .utils import custom_titled_filter
from utilities.date_range_filter import CustomDateRangeFilter, CustomDateTimeRangeFilter
from django.urls import path
from openpyxl import load_workbook
from django.contrib import messages
from django.http import HttpResponseRedirect
from django import forms
from django.shortcuts import resolve_url, render
from instant_cashin.tasks import update_manual_batch_transactions_task


class DistinctFilter(admin.SimpleListFilter):
    title = "Distinct"
    parameter_name = "Distinct"

    def lookups(self, request, model_admin):
        return (("distinct", "Distinct"),)

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        if self.value() == "distinct":
            # return queryset.distinct("parent_transaction__transaction_id").order_by("parent_transaction__transaction_id", "-id")
            return queryset.filter(
                id__in=[
                    trn.id
                    for trn in queryset.distinct(
                        "parent_transaction__transaction_id"
                    ).order_by("parent_transaction__transaction_id", "-id")
                ]
            )


class DisbursedFilter(admin.SimpleListFilter):
    title = "Disbursement Status"
    parameter_name = "disbursed"

    def lookups(self, request, model_admin):
        return (
            ("yes", "Yes"),
            ("no", "No"),
        )

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        if self.value() == "yes":
            return queryset.filter(~Q(disbursed_date=None))
        elif self.value() == "no":
            return queryset.filter(disbursed_date=None, reason="")


class TransactionStatusFilter(admin.SimpleListFilter):
    title = "Transaction Status"
    parameter_name = "status"

    def lookups(self, request, model_admin):
        return (
            ("S", "Successful"),
            ("F", "Failed"),
        )

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        if self.value() == "S":
            return queryset.filter(
                ~Q(disbursed_date=None), ~Q(reason=""), Q(is_disbursed=True)
            )
        elif self.value() == "F":
            return queryset.filter(
                ~Q(disbursed_date=None), ~Q(reason=""), Q(is_disbursed=False)
            )


class EndToEndFilter(admin.SimpleListFilter):
    title = "Has End To End Transaction"
    parameter_name = "end_to_end"

    def lookups(self, request, model_admin):
        return (
            ("yes", "Yes"),
            ("no", "No"),
        )

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        if self.value() == "yes":
            return queryset.filter(~Q(end_to_end=""))
        elif self.value() == "no":
            return queryset.filter(Q(end_to_end=""))


class TimeoutFilter(admin.SimpleListFilter):
    title = "Timeout Status"
    parameter_name = "is_timeout"

    def lookups(self, request, model_admin):
        return (("yes", "Yes"),)

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        if self.value() == "yes":
            return queryset.filter(
                ~Q(disbursed_date=None), reason="", is_disbursed=False
            )


@admin.register(BankTransaction)
class BankTransactionAdminModel(
    admin.ModelAdmin, BankExportCsvMixin, ExportCsvMixin, BankExportExcelMixin
):
    """
    Admin model for customizing BankTransaction model admin view
    """

    list_display = [
        "transaction_id",
        "parent_transaction",
        "creditor_account_number",
        "creditor_bank",
        "category_code",
        "amount",
        "status",
        "transaction_status_code",
        "created_at",
        "disbursed_date",
    ]
    search_fields = [
        "transaction_id",
        "parent_transaction__transaction_id",
        "creditor_account_number",
        "client_transaction_reference",
    ]
    readonly_fields = [field.name for field in BankTransaction._meta.local_fields]
    list_filter = [
        ("disbursed_date", CustomDateTimeRangeFilter),
        ("created_at", CustomDateRangeFilter),
        DistinctFilter,
        EndToEndFilter,
        "is_manual_batch",
        "is_exported_for_manual_batch",
        "status",
        "category_code",
        "transaction_status_code",
        "is_single_step",
        "user_created__root",
    ]
    ordering = ["-id"]
    fieldsets = (
        (
            None,
            {
                "fields": (
                    "parent_transaction",
                    "transaction_id",
                    "message_id",
                    "user_created",
                    "status",
                    "amount",
                    "currency",
                    "purpose",
                    "category_code",
                    "transaction_status_code",
                    "transaction_status_description",
                    "is_single_step",
                    "document",
                    'fees',
                    'vat'
                )
            },
        ),
        (
            _("Creditor/Beneficiary Data"),
            {
                "fields": (
                    "creditor_name",
                    "creditor_account_number",
                    "creditor_bank",
                    "creditor_bank_branch",
                    "creditor_email",
                    "creditor_mobile_number",
                    "creditor_id",
                    "creditor_address_1",
                    "comment",
                )
            },
        ),
        (
            _("Debtor Data"),
            {
                "fields": (
                    "debtor_account",
                    "corporate_code",
                    "sender_id",
                    "debtor_address_1",
                )
            },
        ),
        (
            _("Additional Transaction Info"),
            {"fields": ("additional_info_1", "end_to_end")},
        ),
        (_("Important Dates"), {"fields": ("created_at", "updated_at")}),
        (_("Balance updates"), {"fields": ("balance_before", "balance_after")}),
        (
            _("Manual Batch"),
            {
                "fields": (
                    "is_manual_batch",
                    "is_exported_for_manual_batch",
                    "bank_batch_id",
                    "bank_transaction_id",
                    "bank_end_to_end_identifier",
                )
            },
        ),
    )
    actions = ["export_as_csv", "export_bulk_as_csv", "export_bulk_as_excel"]

    def has_module_permission(self, request):
        if request.user.is_superuser or request.user.has_perm(
            "users.has_instant_transaction_view"
        ) or request.user.is_single_step_support :
            return True

    def has_view_permission(self, request, obj=None):
        if request.user.is_superuser or request.user.has_perm(
            "users.has_instant_transaction_view"
        ) or request.user.is_single_step_support :
            return True

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def get_urls(self):
        urls = super().get_urls()
        new_urls = [
            path("manual-batch-updates/", self.manual_batch_updates),
        ]
        return new_urls + urls

    def manual_batch_updates(self, request):

        if request.method == "POST":
            xlsx_file = request.FILES["manual_batch_file"]

            if not xlsx_file.name.endswith(".xlsx"):
                messages.warning(request, "Wrong file type was uploaded")
                return HttpResponseRedirect(request.path_info)

            wb = load_workbook(xlsx_file)
            ws = wb.active
            last_row = len(list(ws.rows))
            my_dict = []
            for row in range(2, last_row + 1):
                my_dict.append(
                    {
                        "transaction_id": ws["X" + str(row)].value.replace("\n", ""),
                        "bank_batch_id": ws["A" + str(row)].value,
                        "bank_transaction_id": ws["G" + str(row)].value,
                        "bank_end_to_end_identifier": ws["F" + str(row)].value,
                        "amount": ws["H" + str(row)].value,
                        "status": ws["AC" + str(row)].value,
                        "status_decription": ws["AA" + str(row)].value,
                    }
                )
            update_manual_batch_transactions_task.run(my_dict)

        form = ManualBatchUpdateStatusForm()
        data = {"form": form}
        return render(request, "admin/manual_batch.html", data)


class ManualBatchUpdateStatusForm(forms.Form):
    manual_batch_file = forms.FileField()


@admin.register(Agent)
class AgentAdmin(admin.ModelAdmin):
    """
    Admin model for the tweaking the representation of the Agent model at the admin panel
    """

    list_display = ["msisdn", "wallet_provider", "super", "pin", "type"]
    list_filter = ["pin", "super", "wallet_provider", "type"]
    search_fields = ["msisdn"]


@admin.register(DisbursementData)
class DisbursementDataAdmin(
    AdminSiteOwnerOnlyPermissionMixin, admin.ModelAdmin, ExportCsvMixin
):
    """
    Admin panel representation for DisbursementData model
    """

    list_display = [
        "_trx_id",
        "reference_id",
        "msisdn",
        "amount",
        "issuer",
        "is_disbursed",
        "reason",
        "disbursed_date",
    ]
    list_filter = [
        ("disbursed_date", CustomDateRangeFilter),
        ("created_at", CustomDateRangeFilter),
        ("updated_at", CustomDateRangeFilter),
        TransactionStatusFilter,
        DisbursedFilter,
        TimeoutFilter,
        "issuer",
        (
            "doc__file_category__user_created__client__creator",
            custom_titled_filter("Super Admin"),
        ),
        ("doc__file_category__user_created", custom_titled_filter("Entity Admin")),
        ("doc__owner", custom_titled_filter("Document Owner/Uploader")),
    ]
    search_fields = ["id", "doc__id", "msisdn", "reason"]
    ordering = ["-updated_at", "-created_at"]
    actions = ["export_as_csv"]

    fieldsets = (
        (None, {"fields": list_display + ["doc", "_disbursement_document"]}),
        (_("Important Dates"), {"fields": ("created_at", "updated_at")}),
        (_("Balance updates"), {"fields": ("balance_before", "balance_after")}),
    )

    def _trx_id(self, obj):
        """Add transaction id field renamed to Trx ID"""
        return obj.id

    _trx_id.short_description = "Trx ID"

    def _disbursement_document(self, instance):
        """Link to the original disbursement document"""
        url = resolve_url(
            admin_urlname(DisbursementDocData._meta, "change"),
            instance.doc.disbursement_txn.id,
        )
        return format_html(f"<a href='{url}'>{instance.doc.id}</a>")

    _disbursement_document.short_description = "Go to the disbursement document"

    def has_module_permission(self, request):
        if request.user.is_superuser or request.user.is_single_step_support :
            return True

    def has_view_permission(self, request, obj=None):
        if request.user.is_superuser or request.user.is_single_step_support :
            return True
        
    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(DisbursementDocData)
class DisbursementDocDataAdmin(AdminSiteOwnerOnlyPermissionMixin, admin.ModelAdmin):
    """
    Admin panel representation for DisbursementDocData model
    """

    list_display = [
        "doc",
        "doc_status",
        "txn_id",
        "txn_status",
        "has_callback",
        "updated_at",
    ]
    list_filter = [
        "has_callback",
        ("created_at", CustomDateRangeFilter),
        "doc_status",
        ("doc__owner", custom_titled_filter("Document Owner/Uploader")),
    ]
    search_fields = ["doc__file", "doc__id", "txn_id"]
    fieldsets = (
        (
            None,
            {"fields": ("doc", "txn_id", "txn_status", "doc_status", "has_callback")},
        ),
        (_("Important Dates"), {"fields": ("created_at", "updated_at")}),
    )
    def has_module_permission(self, request):
        if request.user.is_superuser or request.user.is_single_step_support :
            return True

    def has_view_permission(self, request, obj=None):
        if request.user.is_superuser or request.user.is_single_step_support :
            return True
        
    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(VMTData)
class VMTDataAdmin(admin.ModelAdmin):
    """
    Admin model for VMTData credentials
    """

    list_filter = ["vmt", "vmt_environment"]
    list_display = list_filter + [
        "wallet_issuer",
        "login_username",
        "login_password",
        "request_gateway_code",
        "request_gateway_type",
    ]

    def has_add_permission(self, request):
        """
        :return: Prevent non superuser from adding new instances of VMTData credentials
        """
        if request.user.is_superuser:
            return True
        try:
            if request.user.root.vmt:
                return False
        except VMTData.DoesNotExist:
            return super(VMTDataAdmin, self).has_add_permission(request)


@admin.register(RemainingAmounts)
class RemainingAmountsAdmin(admin.ModelAdmin):
    list_display = ("mobile", "full_name", "base_amount", "remaining_amount")

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False
