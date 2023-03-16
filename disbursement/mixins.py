# -*- coding: UTF-8 -*-
from __future__ import unicode_literals

import csv

from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponseRedirect
from django.utils.translation import gettext as _

from disbursement.models import BankTransaction
from disbursement.tasks import ExportFromDjangoAdmin
from utilities.models.abstract_models import AbstractBaseACHTransactionStatus


class AdminSiteOwnerOnlyPermissionMixin:
    """
    For handling add/change/delete permission at the admin panel
    """

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        if (
            request.user.is_superuser
            or request.user == obj.doc.owner.root.super_admin
            or request.user == obj.doc.owner.root
        ):
            return True
        raise PermissionError(
            _(
                "Only admin family member users allowed to delete records from this table."
            )
        )

    def has_change_permission(self, request, obj=None):
        return False


class AdminOrCheckerRequiredMixin(LoginRequiredMixin):
    """
    Check if the user accessing resource is admin or checker with disbursement and accept vf on-boarding permissions
    """

    def dispatch(self, request, *args, **kwargs):
        status = self.request.user.get_status(self.request)

        if not (
            status == "disbursement"
            and self.request.user.is_accept_vodafone_onboarding
            and (self.request.user.is_checker or self.request.user.is_root)
        ):
            return self.handle_no_permission()
        return super().dispatch(request, *args, **kwargs)


class AdminOrCheckerOrSupportRequiredMixin(LoginRequiredMixin):
    """
    Check if the user accessing resource is admin or checker or support
    with disbursement and accept vf on-boarding permissions
    """

    def dispatch(self, request, *args, **kwargs):
        if (
            self.request.user.is_support
            and self.request.user.is_accept_vodafone_onboarding
        ):
            return super().dispatch(request, *args, **kwargs)
        else:
            status = self.request.user.get_status(self.request)

            if not (
                status == "disbursement"
                and self.request.user.is_accept_vodafone_onboarding
                and (self.request.user.is_checker or self.request.user.is_root)
            ):
                return self.handle_no_permission()
        return super().dispatch(request, *args, **kwargs)


class ExportCsvMixin:
    """
    mixin class to add  export to any model
    """

    def export_as_csv(self, request, queryset):

        ids_list = list(queryset.values_list('id', flat=True))

        # fire celery task to export data
        ExportFromDjangoAdmin.delay(request.user.id, ids_list, self.model.__name__)

        self.message_user(
            request, f"Exported data will send to your email in a few minutes"
        )
        return HttpResponseRedirect(request.get_full_path())

    export_as_csv.short_description = "Export Selected"


class BankExportCsvMixin:
    """
    mixin class to add  export to any model
    """

    def export_bulk_as_csv(self, request, queryset):

        meta = self.model._meta
        # field_names = [field.name for field in meta.fields]
        field_names = [
            "TransactionID",
            "CreditorAccountNumber",
            "CreditorBank",
            "CreditorBankBranch",
            "TransactionAmount",
            "TransactionPurpose",
            "Comments",
            "ReceiverEmail",
        ]
        _field_names = [
            "transaction_id",
            "creditor_account_number",
            "creditor_bank",
            "creditor_bank_branch",
            "amount",
            "purpose",
            "comment",
        ]
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = "attachment; filename={}.csv".format(meta)
        writer = csv.writer(response)

        writer.writerow(field_names)
        for obj in queryset:
            if queryset.model is BankTransaction:
                obj.status = [
                    st
                    for st in AbstractBaseACHTransactionStatus.STATUS_CHOICES
                    if st[0] == obj.status
                ][0][1]
            row = writer.writerow(
                [
                    getattr(obj, field) if not field == "purpose" else ""
                    for field in _field_names
                ]
            )

        return response

    export_bulk_as_csv.short_description = "Export Bulk Transaction"


from datetime import datetime

from django.http import HttpResponse
from openpyxl import Workbook


class BankExportExcelMixin:
    def export_bulk_as_excel(self, request, queryset):
        bank_queryset = queryset
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename={}{}.xls".format(
            "Payouts bulk ", datetime.now().strftime("%Y-%m-%d")
        )
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Bulk Transaction"
        columns = [
            "TransactionID",
            "CreditorName",
            "CreditorAccountNumber",
            "CreditorBank",
            "CreditorBankBranch",
            "TransactionAmount",
            "TransactionPurpose",
            "Comments",
            "ReceiverEmail",
        ]
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        for recored in bank_queryset:
            row_num += 1
            row = [
                str(recored.transaction_id),
                recored.creditor_name,
                recored.creditor_account_number,
                recored.creditor_bank,
                recored.creditor_bank_branch,
                recored.amount,
                "",
                str(recored.transaction_id),
                "",
            ]
            recored.is_exported_for_manual_batch = True
            recored.save()
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response
