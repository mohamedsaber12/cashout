import json
from unittest.mock import patch
from data.models import Doc,DocReview
from data.models.filecategory import FileCategory
from openpyxl import load_workbook
import pandas as pd

from rest_framework.test import APITestCase
from rest_framework import status
from rest_framework.reverse import reverse as api_reverse
from urllib.parse import urlencode
from oauth2_provider.models import Application
from users.models.admin import SuperAdminUser
from users.models.base_user import User

from users.tests.factories import (
    InstantAPICheckerFactory, AdminUserFactory, SuperAdminUserFactory,
    VMTDataFactory)
from utilities.models import Budget, FeeSetup
from users.models import Client
from disbursement.models import Agent, DisbursementDocData
from instant_cashin.models import AmanTransaction, InstantTransaction



# -*- coding: utf-8 -*-
from datetime import date
import datetime
import os

from django.test import TestCase, Client, RequestFactory, override_settings
from django.urls import reverse
from django.contrib.auth.models import Permission
from disbursement.forms import VMTDataForm
from instant_cashin.models.instant_transactions import InstantTransaction
from rest_framework_expiring_authtoken.models import ExpiringToken
from users.models.instant_api_checker import InstantAPICheckerUser
from users.models.instant_api_viewer import InstantAPIViewerUser
from users.models.root import RootUser
from users.models.support import SupportSetup, SupportUser
from django.utils.timezone import datetime, make_aware, now
from users.tests.factories import (
    SuperAdminUserFactory, AdminUserFactory, VMTDataFactory
)
from users.models import (
    Setup, Client as ClientModel, Brand, EntitySetup, CheckerUser, Levels,
    MakerUser
)
from data.models import Doc, DocReview, FileCategory
from utilities.models import Budget, CallWalletsModerator, FeeSetup, AbstractBaseDocType
from disbursement.models import Agent, BankTransaction, DisbursementData, DisbursementDocData, VMTData
from unittest.mock import patch
from disbursement.api.views import DisburseAPIView

class CurrentRequest(object):
    def __init__(self, user=None):
        self.user=user



class DisburseAPIViewTests(APITestCase):
    def setUp(self):
        super().setUp()
        # create root
        self.super_admin = SuperAdminUserFactory()
        self.vmt_data_obj = VMTDataFactory(vmt=self.super_admin)
        self.root = AdminUserFactory(user_type=3)
        self.root.root = self.root
        self.brand = Brand(mail_subject='')
        self.brand.save()
        self.root.brand = self.brand
        self.root.set_pin('123456')
        self.root.save()
        self.root.user_permissions. \
            add(Permission.objects.get(
                content_type__app_label='users', codename='has_disbursement')
        )
        self.root.user_permissions. \
            add(Permission.objects.get(
                content_type__app_label='users', codename='accept_vodafone_onboarding')
        )
        self.client_user = ClientModel(client=self.root, creator=self.super_admin)
        self.client_user.save()
        self.checker_user = CheckerUser(
            hierarchy=1,
            id=15,
            username='test_checker_user',
            root=self.root,
            user_type=2
        )
        self.checker_user.save()
        self.level = Levels(
            max_amount_can_be_disbursed=1200,
            created=self.root
        )
        self.level.save()
        self.checker_user.level = self.level
        self.checker_user.save()
        self.checker_user.user_permissions. \
            add(Permission.objects.get(
                content_type__app_label='users', codename='has_disbursement')
        )
        self.checker_user.user_permissions. \
            add(Permission.objects.get(
                content_type__app_label='users', codename='accept_vodafone_onboarding')
        )
        self.maker_user = MakerUser(
            hierarchy=1,
            id=14,
            username='test_maker_user',
            email='t@mk.com',
            root=self.root,
            user_type=1
        )
        self.maker_user.save()
        self.budget = Budget(disburser=self.root, current_balance=250)
        self.budget.save()
        fees_setup_bank_wallet = FeeSetup(budget_related=self.budget, issuer='bc',
                                          fee_type='f', fixed_value=20)
        fees_setup_bank_wallet.save()
        # fees_setup_vodafone = FeeSetup(budget_related=self.budget, issuer='vf',
        #                                fee_type='p', percentage_value=2.25)
        # fees_setup_vodafone.save()
        # create auth data
        Application.objects.create(
                client_type=Application.CLIENT_CONFIDENTIAL, authorization_grant_type=Application.GRANT_PASSWORD,
                name=f"{self.checker_user.username} OAuth App", user=self.checker_user
        )
        # get client_secret and client_id
        auth_data = Application.objects.get(user=self.checker_user)
        # get auth_token
        url = api_reverse("users:oauth2_token")
        data = urlencode({
            "client_id": auth_data.client_id,
            "client_secret": auth_data.client_secret,
            "username": self.checker_user.username,
            "password": "fiA#EmkjLBh9VSXy6XvFKxnR9jXt",
            "grant_type": "password"
        })
        # response = self.client.post(url, data, content_type="application/x-www-form-urlencoded")
        # self.access_token = response.json()['access_token']
        
        file_category = FileCategory.objects.create(
            user_created=self.root,
            no_of_reviews_required =1
        )
        self.doc = Doc.objects.create(
            owner=self.maker_user,
            file_category=file_category,
            is_disbursed=False,
            can_be_disbursed=True,
            is_processed=True,
        )
        doc_review = DocReview.objects.create(
            is_ok=True,
            doc=self.doc,
            user_created=self.checker_user,
        )
        disb_data_doc = DisbursementDocData.objects.create(
            doc=self.doc,
            txn_status = "200",
            has_callback = True,
            doc_status = "5"
        )
        self.client = Client()


    def test_post_method(self):
        self.doc.type_of = AbstractBaseDocType.E_WALLETS
        self.doc.save()
        Disbursementdata_success=DisbursementData.objects.create(doc=self.doc, amount=100,msisdn='01010101010',issuer='vodafone')
        Disbursementdata_pending=DisbursementData.objects.create(doc=self.doc, amount=100,msisdn='01010101011',issuer='etisalat')
        Disbursementdata_failed=DisbursementData.objects.create(doc=self.doc,  amount=100,msisdn='01010101012',issuer='aman')
        url = api_reverse("disb_api:disburse")
        # self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        data = {
            "pin": "123456",
            "doc_id": self.doc.id,
            "user": self.checker_user.username
        }
        response = self.client.post(url, data, content_type='application/json')
        self.assertEqual(response.status_code, 200)

    def test_post_method_with_invalid_pin(self):
        self.doc.type_of = AbstractBaseDocType.E_WALLETS
        self.doc.save()
        Disbursementdata_success=DisbursementData.objects.create(doc=self.doc, amount=100,msisdn='01010101010',issuer='vodafone')
        Disbursementdata_pending=DisbursementData.objects.create(doc=self.doc, amount=100,msisdn='01010101011',issuer='etisalat')
        Disbursementdata_failed=DisbursementData.objects.create(doc=self.doc,  amount=100,msisdn='01010101012',issuer='aman')
        url = api_reverse("disb_api:disburse")
        # self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        data = {
            "pin": "123568",
            "doc_id": self.doc.id,
            "user": self.checker_user.username
        }
        response = self.client.post(url, data, content_type='application/json')

        self.assertEqual(response.status_code, 400)

    @patch("disbursement.api.views.DisburseAPIView.determine_disbursement_status", return_value=False)
    def test_post_method_with_disbursement_status_false(self,mocked):
        self.doc.type_of = AbstractBaseDocType.E_WALLETS
        self.doc.save()
        Disbursementdata_success=DisbursementData.objects.create(doc=self.doc, amount=100,msisdn='01010101010',issuer='vodafone')
        Disbursementdata_pending=DisbursementData.objects.create(doc=self.doc, amount=100,msisdn='01010101011',issuer='etisalat')
        Disbursementdata_failed=DisbursementData.objects.create(doc=self.doc,  amount=100,msisdn='01010101012',issuer='aman')
        url = api_reverse("disb_api:disburse")
        # self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        data = {
            "pin": "123456",
            "doc_id": self.doc.id,
            "user": self.checker_user.username
        }
        response = self.client.post(url, data, content_type='application/json')
        self.assertEqual(response.status_code, 424)

    def test_prepare_agents_list(self):
        self.superagent = Agent(msisdn='01021469732', type='V', wallet_provider=self.super_admin, super=True)
        self.superagent.save()
        self.agent = Agent(msisdn='01021469732', type='V', wallet_provider=self.super_admin, )
        self.agent.save()
        self.agent = Agent(msisdn='01121469732', type='E', wallet_provider=self.super_admin, )
        self.agent.save()

        self.assertEqual(DisburseAPIView.prepare_agents_list(self.root, '123456'), ([{'MSISDN': '01021469732', 'PIN': '123456'}], ['01121469732']))

    def test_prepare_agents_list_using_default_vodafone_user(self):
        self.super_admin = SuperAdminUserFactory()
        self.vmt_data_obj = VMTDataFactory(vmt=self.super_admin)
        self.root1 = AdminUserFactory(user_type=3)
        self.root1.root = self.root
        self.brand = Brand(mail_subject='')
        self.brand.save()
        self.root1.brand = self.brand
        self.root1.set_pin('123456')
        self.root1.save()
        self.root1.user_permissions. \
            add(Permission.objects.get(
                content_type__app_label='users', codename='has_disbursement')
        )
        self.client_user = ClientModel(client=self.root1, creator=self.super_admin)
        self.root1.user_permissions. \
            add(Permission.objects.get(
                content_type__app_label='users', codename='vodafone_default_onboarding')
        )
        self.root1.save()
        self.superagent = Agent(msisdn='01021469732', type='V', wallet_provider=self.root1, super=True)
        self.superagent.save()
        # self.agent = Agent(msisdn='01021469732', type='V', wallet_provider=self.super_admin, )
        # self.agent.save()
        self.agent = Agent(msisdn='01121469732', type='P', wallet_provider=self.root1, )
        self.agent.save()

        self.assertEqual(DisburseAPIView.prepare_agents_list(self.root1, '123456'), ([{'MSISDN': '01121469732', 'PIN': '123456'}], []))


    def test_prepare_vodafone_recipients(self):
        self.doc.type_of = AbstractBaseDocType.E_WALLETS
        self.doc.save()
        Disbursementdata_success=DisbursementData.objects.create(doc=self.doc, amount=100,msisdn='01010101010',issuer='vodafone')
        self.set_disbursed_date=DisburseAPIView.set_disbursed_date(self, doc_id=self.doc.id)
        self.assertEqual(DisburseAPIView.prepare_vodafone_recipients(self, doc_id=self.doc.id), ([{"MSISDN":'01010101010', "AMOUNT":100.0, "TXNID":Disbursementdata_success.id}]))

    def test_determine_disbursement_status(self):
        case1=DisburseAPIView.determine_disbursement_status(self, self.checker_user, self.doc, {'TXNSTATUS':'200','BATCH_ID':155},{})
        case2=DisburseAPIView.determine_disbursement_status(self, self.checker_user, self.doc, {'TXNSTATUS':'200', 'TXNID':578},{})
        case3=DisburseAPIView.determine_disbursement_status(self, self.checker_user, self.doc, {'TXNSTATUS':'201', 'TXNID':578},{'TXNSTATUS':'201'})
        self.assertEqual(case1, True)
        self.assertEqual(case2, True)
        self.assertEqual(case3, False)
   
    # @patch("disbursement.api.views.DisburseAPIView.check_balance_before_disbursement.total_doc_amount", return_value=200)
    # def test_check_balance_before_disbursement(self):
    #     self.request = RequestFactory()
    #     Disbursementdata1=DisbursementData.objects.create(doc=self.doc, amount=100,msisdn='01010101010',issuer='vodafone')
    #     Disbursementdata2=DisbursementData.objects.create(doc=self.doc, amount=100,msisdn='01010101011',issuer='etisalat')
    #     case1=DisburseAPIView.check_balance_before_disbursement(self, self.request, self.checker_user, self.doc)
    #     self.asserEqual(DisburseAPIView.check_balance_before_disbursement(self, self.request, self.checker_user, self.doc),True)


    def test_disburse_for_recipients_deposit(self):
        Disbursementdata2=DisbursementData.objects.create(doc=self.doc, amount=100,msisdn='01010101011',issuer='etisalat')
        case1=DisburseAPIView.disburse_for_recipients_deposit('url',{'WALLETISSUER':'VODAFONE', 'TYPE':'DPSTREQ'},CheckerUser.username,{'WALLETISSUER':'VODAFONE', 'TYPE':'DPSTREQ'},Disbursementdata2)
        # self.asserEqual(case1,None)

        # with self.assertRaises(Exception) as exception_text:
        #     case1=DisburseAPIView.disburse_for_recipients_deposit('',{'WALLETISSUER':'VODAFONE', 'TYPE':'DPSTREQ'},CheckerUser.username,{'WALLETISSUER':'VODAFONE', 'TYPE':'DPSTREQ'},Disbursementdata2)

        # self.assertTrue(
        #     "This user has no children'"
        #     in str(exception_text.exception))

    def test_disburse_for_recipients(self):
        Disbursementdata2=DisbursementData.objects.create(doc=self.doc, amount=100,msisdn='01010101011',issuer='etisalat')
        case1=DisburseAPIView.disburse_for_recipients('url',{'WALLETISSUER':'VODAFONE', 'TYPE':'DPSTREQ'},CheckerUser.username,{'WALLETISSUER':'VODAFONE', 'TYPE':'DPSTREQ'},Disbursementdata2)
        # self.asserEqual(case1,None)

class TestDisburseCallBack(APITestCase):
    def setUp(self):
        super().setUp()
        # create root
        self.root = AdminUserFactory(user_type=3)
        self.root.root = self.root
        self.root.save()
        # create budget
        self.budget = Budget(disburser=self.root, current_balance=100)
        self.budget.save()
        # create api checker
        self.api_checker = InstantAPICheckerFactory(user_type=6, root=self.root)
        # create auth data
        Application.objects.create(
                client_type=Application.CLIENT_CONFIDENTIAL, authorization_grant_type=Application.GRANT_PASSWORD,
                name=f"{self.api_checker.username} OAuth App", user=self.api_checker
        )
        # set password for api checker
        self.api_checker.set_password('fiA#EmkjLBh9VSXy6XvFKxnR9jXt')
        self.api_checker.save()
        # get client_secret and client_id
        auth_data = Application.objects.get(user=self.api_checker)
        # get auth_token
        url = api_reverse("users:oauth2_token")
        data = urlencode({
            "client_id": auth_data.client_id,
            "client_secret": auth_data.client_secret,
            "username": self.api_checker.username,
            "password": "fiA#EmkjLBh9VSXy6XvFKxnR9jXt",
            "grant_type": "password"
        })
        response = self.client.post(url, data, content_type="application/x-www-form-urlencoded")
        self.access_token = response.json()['access_token']
        self.super_admin = SuperAdminUserFactory()
        self.vmt_data_obj = VMTDataFactory(vmt=self.super_admin)

    def test_update(self):
        url = api_reverse("disb_api:disburse_callback")
        self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        response = self.client.put(url, {}, content_type='application/json')





class TestOnboardMerchant(APITestCase):
    def setUp(self):
        # self.super_admin = User.objects.create(username='ACCEPT_VODAFONE_INTERNAL_SUPERADMIN',email='ACCEPT_VODAFONE_INTERNAL_SUPERADMIN@paymob.com')
        # os.environ.setdefault('ACCEPT_VODAFONE_INTERNAL_SUPERADMIN', 'accept_vodafone_internal_super_admin')
        self.super_admin = SuperAdminUser.objects.create(username='accept_vodafone_internal_super_admin',email='ACCEPT_VODAFONE_INTERNAL_SUPERADMIN@paymob.com')
        self.Systemuser=User.objects.create(username="systemuser",email='systemuser@paymob.com',user_type=14)
        self.Systemuser.set_password('fiA#EmkjLBh9VSXy6XvFKxnR9jXt')
        self.Systemuser.save()
        Application.objects.create(
                client_type=Application.CLIENT_CONFIDENTIAL, authorization_grant_type=Application.GRANT_PASSWORD,
                name=f"{self.Systemuser.username} OAuth App", user=self.Systemuser
        )
        # get client_secret and client_id
        auth_data = Application.objects.get(user=self.Systemuser)
        self.url = api_reverse("users:oauth2_token")
        self.data = urlencode({
            "client_id": auth_data.client_id,
            "client_secret": auth_data.client_secret,
            "username": self.Systemuser.username,
            "password": "fiA#EmkjLBh9VSXy6XvFKxnR9jXt",
            "grant_type": "password"
        })
        response = self.client.post(self.url, self.data, content_type="application/x-www-form-urlencoded")
        self.access_token = response.json()['access_token']

    def test_post_method(self):
        body={
            'username':"test",
            'mobile_number':'01010101010',
            # 'email':'test@paymob.com',
            'idms_user_id':'test1'


        
        }
        
        self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        url = api_reverse("disb_api:merchant_creation")
        response = self.client.post(url, json.dumps(body), content_type='application/json')
        self.assertEqual(response.status_code, 201)


    def test_post_method_with_user_not_systemuser(self):

        self.super_admin.set_password('fiA#EmkjLBh9VSXy6XvFKxnR9jXt')
        self.super_admin.save()
        Application.objects.create(
                client_type=Application.CLIENT_CONFIDENTIAL, authorization_grant_type=Application.GRANT_PASSWORD,
                name=f"{self.super_admin.username} OAuth App", user=self.super_admin
        )
        # get client_secret and client_id
        auth_data = Application.objects.get(user=self.super_admin)
        self.url = api_reverse("users:oauth2_token")
        self.data = urlencode({
            "client_id": auth_data.client_id,
            "client_secret": auth_data.client_secret,
            "username": self.super_admin.username,
            "password": "fiA#EmkjLBh9VSXy6XvFKxnR9jXt",
            "grant_type": "password"
        })
        response = self.client.post(self.url, self.data, content_type="application/x-www-form-urlencoded")
        self.access_token = response.json()['access_token']
        body={
            'username':"test",
            'mobile_number':'01010101010',
            'email':'test@paymob.com',
            'idms_user_id':'test1'
        }
        
        self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        url = api_reverse("disb_api:merchant_creation")
        response = self.client.post(url, json.dumps(body), content_type='application/json')
        self.assertEqual(response.status_code, 400)

    def test_post_method_with_user_exists(self):
        self.user = User.objects.create(username='test',email='test1@paymob.com')

        body={
            'username':"test",
            'mobile_number':'01010101010',
            'email':'test@paymob.com',
            'idms_user_id':'test1'


        
        }
        
        self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        url = api_reverse("disb_api:merchant_creation")
        response = self.client.post(url, json.dumps(body), content_type='application/json')
        self.assertEqual(response.status_code, 201)

    def test_post_method_with_error(self):
        self.user = User.objects.create(username='test',email='test@paymob.com',idms_user_id='test1' )

        body={
            'username':"test1",
            'mobile_number':'01010101010',
            'email':'test@paymob.com',
            'idms_user_id':'test1'
        }
        
        self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        url = api_reverse("disb_api:merchant_creation")
        response = self.client.post(url, json.dumps(body), content_type='application/json')
        self.assertEqual(response.status_code, 400)

    def test_post_method_with_exception(self):
        self.super_admin.username='accept_vodafone_internal_super_admin1'
        self.super_admin.save()

        body={
            'username':"test1",
            'mobile_number':'01010101010',
            'email':'test@paymob.com',
            'idms_user_id':'test1'
        }
        
        self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        url = api_reverse("disb_api:merchant_creation")
        response = self.client.post(url, json.dumps(body), content_type='application/json')
        self.assertEqual(response.status_code, 400)


class TestSendMailForCreationAdmin(APITestCase):

    def setUp(self):
        
        self.super_admin = SuperAdminUser.objects.create(username='accept_vodafone_internal_super_admin',email='ACCEPT_VODAFONE_INTERNAL_SUPERADMIN@paymob.com')
        self.Systemuser=User.objects.create(username="systemuser",email='systemuser@paymob.com',user_type=14)
        self.Systemuser.set_password('fiA#EmkjLBh9VSXy6XvFKxnR9jXt')
        self.Systemuser.save()
        Application.objects.create(
                client_type=Application.CLIENT_CONFIDENTIAL, authorization_grant_type=Application.GRANT_PASSWORD,
                name=f"{self.Systemuser.username} OAuth App", user=self.Systemuser
        )
        # get client_secret and client_id
        auth_data = Application.objects.get(user=self.Systemuser)
        self.url = api_reverse("users:oauth2_token")
        self.data = urlencode({
            "client_id": auth_data.client_id,
            "client_secret": auth_data.client_secret,
            "username": self.Systemuser.username,
            "password": "fiA#EmkjLBh9VSXy6XvFKxnR9jXt",
            "grant_type": "password"
        })
        response = self.client.post(self.url, self.data, content_type="application/x-www-form-urlencoded")
        self.access_token = response.json()['access_token']

    def test_post_method(self):
        body={
            'username':"test",
            'mobile_number':'01010101010',
            'email':'test@paymob.com',
            'idms_user_id':'test1',
            'mid':'55525'


        
        }
        
        self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        url = api_reverse("disb_api:send_mail_creation")
        response = self.client.post(url, json.dumps(body), content_type='application/json')
        self.assertEqual(response.status_code, 201)

    def test_post_method_with_user_not_systemuser(self):

        self.super_admin.set_password('fiA#EmkjLBh9VSXy6XvFKxnR9jXt')
        self.super_admin.save()
        Application.objects.create(
                client_type=Application.CLIENT_CONFIDENTIAL, authorization_grant_type=Application.GRANT_PASSWORD,
                name=f"{self.super_admin.username} OAuth App", user=self.super_admin
        )
        # get client_secret and client_id
        auth_data = Application.objects.get(user=self.super_admin)
        self.url = api_reverse("users:oauth2_token")
        self.data = urlencode({
            "client_id": auth_data.client_id,
            "client_secret": auth_data.client_secret,
            "username": self.super_admin.username,
            "password": "fiA#EmkjLBh9VSXy6XvFKxnR9jXt",
            "grant_type": "password"
        })
        response = self.client.post(self.url, self.data, content_type="application/x-www-form-urlencoded")
        self.access_token = response.json()['access_token']
        body={
            'username':"test",
            'mobile_number':'01010101010',
            'email':'test@paymob.com',
            'idms_user_id':'test1',
            'mid':'55525'


        
        }
        
        self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        url = api_reverse("disb_api:send_mail_creation")
        response = self.client.post(url, json.dumps(body), content_type='application/json')
        self.assertEqual(response.status_code, 403)


    def test_post_method_with_error(self):
        self.user = User.objects.create(username='test',email='test@paymob.com',idms_user_id='test1' )

        body={
            'username':"test",
            'mobile_number':'01010101010',
            'email':'test@paymob.com',
            'idms_user_id':'test1',
            # 'mid':'55525'


        
        }
        
        self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        url = api_reverse("disb_api:send_mail_creation")
        response = self.client.post(url, json.dumps(body), content_type='application/json')
        self.assertEqual(response.status_code, 400)


class MockResponse:
 
    def __init__(self):
        self.status_code = 200
        self.text='test'
        self.ok=True
    def json(self):
        return {'TXNSTATUS':'400'}
class TestCreateSingleStepTransacton(APITestCase):

    def setUp(self):
        
        self.super_admin = SuperAdminUser.objects.create(username='accept_vodafone_internal_super_admin',email='ACCEPT_VODAFONE_INTERNAL_SUPERADMIN@paymob.com')
        self.Systemuser=User.objects.create(username="systemuser",email='systemuser@paymob.com',user_type=14)
        self.Systemuser.set_password('fiA#EmkjLBh9VSXy6XvFKxnR9jXt')
        self.Systemuser.save()
        Application.objects.create(
                client_type=Application.CLIENT_CONFIDENTIAL, authorization_grant_type=Application.GRANT_PASSWORD,
                name=f"{self.Systemuser.username} OAuth App", user=self.Systemuser
        )
        # get client_secret and client_id
        auth_data = Application.objects.get(user=self.Systemuser)
        self.url = api_reverse("users:oauth2_token")
        self.data = urlencode({
            "client_id": auth_data.client_id,
            "client_secret": auth_data.client_secret,
            "username": self.Systemuser.username,
            "password": "fiA#EmkjLBh9VSXy6XvFKxnR9jXt",
            "grant_type": "password"
        })
        response = self.client.post(self.url, self.data, content_type="application/x-www-form-urlencoded")
        self.access_token = response.json()['access_token']

    def test_post_method(self):
        body={
            'amount':"100",
            'issuer':'vodafone',
            'username':'test1',
            'msisdn':'01060108097',
            'admin_email':'test@paymob.com',
            'idms_user_id':'ssd5'        
        }
        
        self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        url = api_reverse("disb_api:create_single_step_transaction")
        response = self.client.post(url, json.dumps(body), content_type='application/json')
        self.assertEqual(response.status_code, 400)

    def test_post_method_with_issuer_bankcard(self):
        body={
            'amount':"100",
            'issuer':'bank_card',
            'creditor_name':'test',
            'creditor_account_number':'555555',
            'creditor_bank':'AUB',
            'transaction_type':'SALARY',
            'username':'test1',
            'full_name':'test test',
            'admin_email':'test@paymob.com',
            'idms_user_id':'ssd5' ,
            'msisdn':'545655656565'       
        }
        
        self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        url = api_reverse("disb_api:create_single_step_transaction")
        response = self.client.post(url, json.dumps(body), content_type='application/json')
        self.assertEqual(response.status_code, 400)

    def test_post_method_with_issuer_aman(self):
        body={
            'amount':"100",
            'issuer':'aman',
            'username':'test1',
            'full_name':'test test',
            'admin_email':'test@paymob.com',
            'idms_user_id':'ssd5' ,
            'msisdn':'545655656565',
            'first_name':'first_name',
            'last_name':'last_name',
            'email':'test@paymob.com',       
        }
        
        self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        url = api_reverse("disb_api:create_single_step_transaction")
        response = self.client.post(url, json.dumps(body), content_type='application/json')
        self.assertEqual(response.status_code, 400)

    def test_post_method_with_issuer_bank_wallet(self):
        body={
            'amount':"100",
            'issuer':'bank_wallet',
            'username':'test1',
            'full_name':'test test',
            'admin_email':'test@paymob.com',
            'idms_user_id':'ssd5' ,
            'msisdn':'545655656565',       
        }
        
        self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        url = api_reverse("disb_api:create_single_step_transaction")
        response = self.client.post(url, json.dumps(body), content_type='application/json')
        self.assertEqual(response.status_code, 400)

    def test_post_method_with_user_not_systemuser(self):

        self.super_admin.set_password('fiA#EmkjLBh9VSXy6XvFKxnR9jXt')
        self.super_admin.save()
        Application.objects.create(
                client_type=Application.CLIENT_CONFIDENTIAL, authorization_grant_type=Application.GRANT_PASSWORD,
                name=f"{self.super_admin.username} OAuth App", user=self.super_admin
        )
        # get client_secret and client_id
        auth_data = Application.objects.get(user=self.super_admin)
        self.url = api_reverse("users:oauth2_token")
        self.data = urlencode({
            "client_id": auth_data.client_id,
            "client_secret": auth_data.client_secret,
            "username": self.super_admin.username,
            "password": "fiA#EmkjLBh9VSXy6XvFKxnR9jXt",
            "grant_type": "password"
        })
        response = self.client.post(self.url, self.data, content_type="application/x-www-form-urlencoded")
        self.access_token = response.json()['access_token']
        body={
            'amount':"100",
            'issuer':'vodafone',
            'username':'test1',
            'msisdn':'01060108097',
            'admin_email':'test@paymob.com',
            'idms_user_id':'ssd5'
        }
        
        self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        url = api_reverse("disb_api:create_single_step_transaction")
        response = self.client.post(url, json.dumps(body), content_type='application/json')
        self.assertEqual(response.status_code, 400)

    def test_post_method_with_error(self):
        # self.user = User.objects.create(username='test',email='test@paymob.com',idms_user_id='test1' )

        body={
            # 'amount':"100",
            # 'issuer':'vodafone',
            # 'username':'test1',
            'msisdn':'01060108097',
            'admin_email':'test@paymob.com',
            'idms_user_id':'ssd5'        
        }
        
        self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        url = api_reverse("disb_api:create_single_step_transaction")
        response = self.client.post(url, json.dumps(body), content_type='application/json')
        self.assertEqual(response.status_code, 400)

    def test_post_method_with_exception(self):
        self.super_admin.username='accept_vodafone_internal_super_admin1'
        self.super_admin.save()

        body={
            'username':"test1",
            'mobile_number':'01010101010',
            'email':'test@paymob.com',
            'idms_user_id':'test1'
        }
        
        self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        url = api_reverse("disb_api:create_single_step_transaction")
        response = self.client.post(url, json.dumps(body), content_type='application/json')
        self.assertEqual(response.status_code, 400)


    @patch("requests.post", return_value=MockResponse())
    def test_post_method_with_issuer_vodafone(self, mocked):
        body={
            'amount':"100",
            'issuer':'vodafone',
            'username':'test1',
            'msisdn':'01060108097',
            'admin_email':'test@paymob.com',
            'idms_user_id':'ssd5'        
        }
        
        self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        url = api_reverse("disb_api:create_single_step_transaction")
        response = self.client.post(url, json.dumps(body), content_type='application/json')
        self.assertEqual(response.status_code, 201)

class TestRetrieveDocData(APITestCase):
    def setUp(self):
        self.now = datetime.now()
        self.super_admin = SuperAdminUserFactory()
        self.vmt_data_obj = VMTDataFactory(vmt=self.super_admin)
        self.root = AdminUserFactory(user_type=3)
        self.root.root = self.root
        self.brand = Brand(mail_subject='')
        self.brand.save()
        self.root.brand = self.brand
        self.root.set_pin('123456')
        self.root.save()
        self.root.user_permissions. \
            add(Permission.objects.get(
                content_type__app_label='users', codename='has_disbursement'
        )
        )
        self.root.user_permissions. \
            add(Permission.objects.get(
                content_type__app_label='users', codename='accept_vodafone_onboarding'
        )
        )
        self.client_user = ClientModel(client=self.root, creator=self.super_admin)
        self.client_user.save()
        self.setup = Setup.objects.create(
                user=self.root,
                levels_setup=True,
                maker_setup=True,
                checker_setup=True,
                category_setup=True,
                pin_setup=True
        )
        self.entity_setup = EntitySetup.objects.create(
                user=self.super_admin,
                entity=self.root,
                agents_setup=True,
                fees_setup=True
        )
        self.checker_user = CheckerUser(
                id=15,
                username='test_checker_user',
                root=self.root,
                user_type=2
        )
        self.checker_user.save()
        self.level = Levels(
                max_amount_can_be_disbursed=1200,
                created=self.root
        )
        self.level.save()
        self.checker_user.level = self.level
        self.checker_user.save()
        self.checker_user.user_permissions. \
            add(Permission.objects.get(
                content_type__app_label='users', codename='has_disbursement'
        )
        )
        self.checker_user.user_permissions. \
            add(Permission.objects.get(
                content_type__app_label='users', codename='accept_vodafone_onboarding'
        )
        )
        self.maker_user = MakerUser(
                id=14,
                username='test_maker_user',
                email='t@mk.com',
                root=self.root,
                user_type=1
        )
        self.maker_user.save()
        self.budget = Budget(disburser=self.root, current_balance=150)
        self.budget.save()
        fees_setup_bank_wallet = FeeSetup(budget_related=self.budget, issuer='bc',
                                          fee_type='f', fixed_value=20)
        fees_setup_bank_wallet.save()
        fees_setup_vodafone = FeeSetup(budget_related=self.budget, issuer='vf',
                                       fee_type='p', percentage_value=2.25)
        fees_setup_vodafone.save()

        # create doc, doc_review, DisbursementDocData, file category
        self.file_category = FileCategory.objects.create(
                user_created=self.root
        )
        # self.doc = Doc.objects.create(
        #         owner=self.maker_user,
        #         file_category=file_category,
        #         is_disbursed=True,
        #         can_be_disbursed=True,
        #         is_processed=True,
        #         # file=f'documents/files_uploaded/{self.now.year}/{self.now.day}/bank_wallets_sample_file_{self.maker_user.username}_72RYZL00Y.xlsx'
        # )
        self.checker_user.set_password('fiA#EmkjLBh9VSXy6XvFKxnR9jXt')
        self.checker_user.save()
        Application.objects.create(
                client_type=Application.CLIENT_CONFIDENTIAL, authorization_grant_type=Application.GRANT_PASSWORD,
                name=f"{self.checker_user.username} OAuth App", user=self.checker_user
        )
        # get client_secret and client_id
        auth_data = Application.objects.get(user=self.checker_user)
        self.url = api_reverse("users:oauth2_token")
        self.data = urlencode({
            "client_id": auth_data.client_id,
            "client_secret": auth_data.client_secret,
            "username": self.checker_user.username,
            "password": "fiA#EmkjLBh9VSXy6XvFKxnR9jXt",
            "grant_type": "password"
        })
        response = self.client.post(self.url, self.data, content_type="application/x-www-form-urlencoded")
        self.access_token = response.json()['access_token']

    def test_get_method(self):
        
        if not os.path.exists(f'/app/mediafiles/media/documents/files_uploaded/{self.now.year}/{self.now.day}'):
            os.makedirs(f'/app/mediafiles/media/documents/files_uploaded/{self.now.year}/{self.now.day}')
        # with open(f'/app/mediafiles/media/documents/files_uploaded/{self.now.year}/{self.now.day}/bank_wallets_sample_file_{self.maker_user.username}_72RYZL00Y.xls', 'w') as fp:
        #     fp.write("file information")
        #     fp.close()
        #     pass
        # wb= load_workbook(f'/app/mediafiles/media/documents/files_uploaded/{self.now.year}/{self.now.day}/bank_wallets_sample_file_{self.maker_user.username}_72RYZL00Y.xls')
        # sheets=wb.sheetnames
        # sheet1= wb[sheets[1]]
        # sheet1.cell(row=, column=).value=

        df=pd.DataFrame(
            {
                'mobile number':['01211409281'],
                'amount':['55'],
                'issuer':['vodafone']
            }
        )
        with pd.ExcelWriter(f'/app/mediafiles/media/documents/files_uploaded/{self.now.year}/{self.now.day}/bank_wallets_sample_file_{self.maker_user.username}_72RYZL00Y.xls') as writer:
            df.to_excel(writer, sheet_name='sheet1')

        self.doc = Doc.objects.create(
                owner=self.maker_user,
                file_category=self.file_category,
                is_disbursed=True,
                can_be_disbursed=True,
                is_processed=True,
                file=f'documents/files_uploaded/{self.now.year}/{self.now.day}/bank_wallets_sample_file_{self.maker_user.username}_72RYZL00Y.xls'
        )
        self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        url = api_reverse("disb_api:docrows", kwargs={'doc_id':self.doc.id})
        response = self.client.get(url, content_type='application/json')
        self.assertEqual(response.status_code, 404)

    def test_get_method_with_doc_not_found(self):
        

        self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        url = api_reverse("disb_api:docrows", kwargs={'doc_id':11})
        response = self.client.get(url, content_type='application/json')
        self.assertEqual(response.status_code, 404)



class TestAllowDocDisburse(APITestCase):

    def setUp(self):
        self.super_admin = SuperAdminUserFactory()
        self.vmt_data_obj = VMTDataFactory(vmt=self.super_admin)
        self.root = AdminUserFactory(user_type=3)
        self.root.root = self.root
        self.brand = Brand(mail_subject='')
        self.brand.save()
        self.root.brand = self.brand
        self.root.set_pin('123456')
        self.root.save()
        self.root.user_permissions. \
            add(Permission.objects.get(
                content_type__app_label='users', codename='has_disbursement'
        )
        )
        self.root.user_permissions. \
            add(Permission.objects.get(
                content_type__app_label='users', codename='accept_vodafone_onboarding'
        )
        )
        self.client_user = ClientModel(client=self.root, creator=self.super_admin)
        self.client_user.save()
        self.setup = Setup.objects.create(
                user=self.root,
                levels_setup=True,
                maker_setup=True,
                checker_setup=True,
                category_setup=True,
                pin_setup=True
        )
        self.entity_setup = EntitySetup.objects.create(
                user=self.super_admin,
                entity=self.root,
                agents_setup=True,
                fees_setup=True
        )
        self.checker_user = CheckerUser(
                id=15,
                username='test_checker_user',
                root=self.root,
                user_type=2
        )
        self.checker_user.save()
        # self.level = Levels(
        #         max_amount_can_be_disbursed=1200,
        #         created=self.root
        # )
        # self.level.save()
        # self.checker_user.level = self.level
        # self.checker_user.save()
        self.checker_user.user_permissions. \
            add(Permission.objects.get(
                content_type__app_label='users', codename='has_disbursement'
        )
        )
        self.checker_user.user_permissions. \
            add(Permission.objects.get(
                content_type__app_label='users', codename='accept_vodafone_onboarding'
        )
        )
        self.maker_user = MakerUser(
                id=14,
                username='test_maker_user',
                email='t@mk.com',
                root=self.root,
                user_type=1
        )
        self.maker_user.save()
        self.budget = Budget(disburser=self.root, current_balance=150)
        self.budget.save()
        fees_setup_bank_wallet = FeeSetup(budget_related=self.budget, issuer='bc',
                                          fee_type='f', fixed_value=20)
        fees_setup_bank_wallet.save()
        fees_setup_vodafone = FeeSetup(budget_related=self.budget, issuer='vf',
                                       fee_type='p', percentage_value=2.25)
        fees_setup_vodafone.save()

        # create doc, doc_review, DisbursementDocData, file category
        self.file_category = FileCategory.objects.create(
                user_created=self.root
        )
        # self.doc = Doc.objects.create(
        #         owner=self.maker_user,
        #         file_category=self.file_category,
        #         is_disbursed=True,
        #         can_be_disbursed=True,
        #         is_processed=True,
        # )
        self.maker_user.set_password('fiA#EmkjLBh9VSXy6XvFKxnR9jXt')
        self.maker_user.save()
        Application.objects.create(
                client_type=Application.CLIENT_CONFIDENTIAL, authorization_grant_type=Application.GRANT_PASSWORD,
                name=f"{self.maker_user.username} OAuth App", user=self.maker_user
        )
        # get client_secret and client_id
        auth_data = Application.objects.get(user=self.maker_user)
        self.url = api_reverse("users:oauth2_token")
        self.data = urlencode({
            "client_id": auth_data.client_id,
            "client_secret": auth_data.client_secret,
            "username": self.maker_user.username,
            "password": "fiA#EmkjLBh9VSXy6XvFKxnR9jXt",
            "grant_type": "password"
        })
        response = self.client.post(self.url, self.data, content_type="application/x-www-form-urlencoded")
        self.access_token = response.json()['access_token']

    def test_get_method_with_doc_can_be_disbursed(self):

        self.doc = Doc.objects.create(
                owner=self.maker_user,
                file_category=self.file_category,
                is_disbursed=True,
                can_be_disbursed=True,
                is_processed=True,
        )
        self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        url = api_reverse("disb_api:allow_doc_disburse", kwargs={'doc_id':self.doc.id})
        response = self.client.post(url, content_type='application/json')
        self.assertEqual(response.status_code, 400)




    def test_get_method(self):

        self.level = Levels(
                max_amount_can_be_disbursed=1200,
                created=self.root
        )
        self.level.save()
        self.checker_user.level = self.level
        self.checker_user.save()
        self.doc = Doc.objects.create(
                owner=self.maker_user,
                file_category=self.file_category,
                is_disbursed=True,
                can_be_disbursed=False,
                is_processed=True,
        )
        self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        url = api_reverse("disb_api:allow_doc_disburse", kwargs={'doc_id':self.doc.id})
        response = self.client.post(url, content_type='application/json')
        self.assertEqual(response.status_code, 200)


class Mock_void_transaction:
 
    def __init__(self):
        self.status_code = 200
        self.text='test'
        self.ok=True
    def json(self):
        return {"success":True}

# def Mock_void_transaction():
#     transactions={
#     "type":"TRANSACTION",
#     "obj":{
#       "id":78,
#       "pending": "false",
#       "amount_cents":100,
#       "success":"true",}
#       }
#     return transactions
class TestCancelAmanTransactionView(APITestCase):

    def setUp(self):
        super().setUp()
        # create root
        self.client = Client()

        self.super_admin = SuperAdminUserFactory()
        self.vmt_data_obj = VMTDataFactory(vmt=self.super_admin)
        self.root = AdminUserFactory(user_type=3)
        self.root.root = self.root
        self.brand = Brand(mail_subject='')
        self.brand.save()
        self.root.brand = self.brand
        self.root.set_pin('123456')
        self.root.save()
        self.root.user_permissions. \
            add(Permission.objects.get(
                content_type__app_label='users', codename='has_disbursement')
        )
        self.root.user_permissions. \
            add(Permission.objects.get(
                content_type__app_label='users', codename='accept_vodafone_onboarding')
        )
        self.client_user = ClientModel(client=self.root, creator=self.super_admin)
        self.client_user.save()
        self.checker_user = CheckerUser(
            hierarchy=1,
            id=15,
            username='test_checker_user',
            root=self.root,
            user_type=2
        )
        self.checker_user.save()
        self.level = Levels(
            max_amount_can_be_disbursed=1200,
            created=self.root
        )
        self.level.save()
        self.checker_user.level = self.level
        self.checker_user.save()
        self.checker_user.user_permissions. \
            add(Permission.objects.get(
                content_type__app_label='users', codename='has_disbursement')
        )
        self.checker_user.user_permissions. \
            add(Permission.objects.get(
                content_type__app_label='users', codename='accept_vodafone_onboarding')
        )
        self.maker_user = MakerUser(
            hierarchy=1,
            id=14,
            username='test_maker_user',
            email='t@mk.com',
            root=self.root,
            user_type=1
        )
        self.maker_user.save()
        self.budget = Budget(disburser=self.root, current_balance=250)
        self.budget.save()
        fees_setup_bank_wallet = FeeSetup(budget_related=self.budget, issuer='bc',
                                          fee_type='f', fixed_value=20)
        fees_setup_bank_wallet.save()
        # fees_setup_vodafone = FeeSetup(budget_related=self.budget, issuer='vf',
        #                                fee_type='p', percentage_value=2.25)
        # fees_setup_vodafone.save()
        # create auth data
        self.checker_user.set_password('fiA#EmkjLBh9VSXy6XvFKxnR9jXt')
        self.checker_user.save()
        Application.objects.create(
                client_type=Application.CLIENT_CONFIDENTIAL, authorization_grant_type=Application.GRANT_PASSWORD,
                name=f"{self.checker_user.username} OAuth App", user=self.checker_user
        )
        # get client_secret and client_id
        auth_data = Application.objects.get(user=self.checker_user)
        # get auth_token
        url = api_reverse("users:oauth2_token")
        data = urlencode({
            "client_id": auth_data.client_id,
            "client_secret": auth_data.client_secret,
            "username": self.checker_user.username,
            "password": "fiA#EmkjLBh9VSXy6XvFKxnR9jXt",
            "grant_type": "password"
        })
        response = self.client.post(url, data, content_type="application/x-www-form-urlencoded")
        self.access_token = response.json()['access_token']
        
        file_category = FileCategory.objects.create(
            user_created=self.root,
            no_of_reviews_required =1
        )
        self.doc = Doc.objects.create(
            owner=self.maker_user,
            file_category=file_category,
            is_disbursed=False,
            can_be_disbursed=True,
            is_processed=True,
        )
        doc_review = DocReview.objects.create(
            is_ok=True,
            doc=self.doc,
            user_created=self.checker_user,
        )
        disb_data_doc = DisbursementDocData.objects.create(
            doc=self.doc,
            txn_status = "200",
            has_callback = True,
            doc_status = "5"
        )


    def test_post_method(self):
        self.doc.type_of = AbstractBaseDocType.E_WALLETS
        self.doc.save()
        Disbursementdata_success=DisbursementData.objects.create(doc=self.doc, amount=100,msisdn='01010101010',issuer='aman')
      
        url = api_reverse("disb_api:cancel_aman_transaction")
        # self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        data = {
           
            'transaction_id':Disbursementdata_success.id
        }

        header={
            'HTTP_X_REQUESTED_WITH':'XMLHttpRequest',
            'HTTP_AUTHORIZATION':'Bearer ' + self.access_token
        }
        response = self.client.post(url, content_type='application/json',data=data,**header)
        # response = self.client.post(reverse("disb_api:cancel_aman_transaction"), data, **{'HTTP_X_REQUESTED_WITH': 'XMLHttpRequest'})

     
        self.assertEqual(response.status_code, 200)
    
    @patch("disbursement.api.views.CancelAmanTransactionView.void_transaction", return_value=Mock_void_transaction())
    def test_post_method_with_mock_void_transaction(self, mocked):
        self.doc.type_of = AbstractBaseDocType.E_WALLETS
        self.doc.save()
        Disbursementdata_success=DisbursementData.objects.create(doc=self.doc, amount=100,msisdn='01010101010',issuer='aman',)
        Disbursementdata_success.reference_id= Disbursementdata_success.id
        Disbursementdata_success.save()
        AmanTransaction.objects.create(transaction=Disbursementdata_success, bill_reference=5455)

      
        url = api_reverse("disb_api:cancel_aman_transaction")
        # self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        data = {
           
            'transaction_id':Disbursementdata_success.id
        }

        header={
            'HTTP_X_REQUESTED_WITH':'XMLHttpRequest',
            'HTTP_AUTHORIZATION':'Bearer ' + self.access_token
        }
        response = self.client.post(url, content_type='application/json',data=data,**header)
        # response = self.client.post(reverse("disb_api:cancel_aman_transaction"), data, **{'HTTP_X_REQUESTED_WITH': 'XMLHttpRequest'})

     
        self.assertEqual(response.status_code, 200)

    def test_post_method_but_not_ajax(self):
        self.doc.type_of = AbstractBaseDocType.E_WALLETS
        self.doc.save()
        Disbursementdata_success=DisbursementData.objects.create(doc=self.doc, amount=100,msisdn='01010101010',issuer='aman')
      
        url = api_reverse("disb_api:cancel_aman_transaction")
        # self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        data = {
           
            'transaction_id':Disbursementdata_success.id
        }

        header={
            # 'HTTP_X_REQUESTED_WITH':'XMLHttpRequest',
            'HTTP_AUTHORIZATION':'Bearer ' + self.access_token
        }
        response = self.client.post(url, content_type='application/json',data=data,**header)
        # response = self.client.post(reverse("disb_api:cancel_aman_transaction"), data, **{'HTTP_X_REQUESTED_WITH': 'XMLHttpRequest'})

     
        self.assertEqual(response.status_code, 403)

    def test_post_method_Without_data(self):
        self.doc.type_of = AbstractBaseDocType.E_WALLETS
        self.doc.save()
        Disbursementdata_success=DisbursementData.objects.create(doc=self.doc, amount=100,msisdn='01010101010',issuer='aman')
      
        url = api_reverse("disb_api:cancel_aman_transaction")
        # self.client.credentials(HTTP_AUTHORIZATION='Bearer ' + self.access_token)
        data = {
           
            # 'transaction_id':Disbursementdata_success.id
        }

        header={
            'HTTP_X_REQUESTED_WITH':'XMLHttpRequest',
            'HTTP_AUTHORIZATION':'Bearer ' + self.access_token
        }
        response = self.client.post(url, content_type='application/json',data=data,**header)
        # response = self.client.post(reverse("disb_api:cancel_aman_transaction"), data, **{'HTTP_X_REQUESTED_WITH': 'XMLHttpRequest'})

     
        self.assertEqual(response.status_code, 401)


# class TestDisbursementDataViewSet(APITestCase):

#     def setUp(self):
#         super().setUp()
#         # create root
#         self.client = Client()

#         self.super_admin = SuperAdminUserFactory()
#         self.vmt_data_obj = VMTDataFactory(vmt=self.super_admin)
#         self.root = AdminUserFactory(user_type=3)
#         self.root.root = self.root
#         self.brand = Brand(mail_subject='')
#         self.brand.save()
#         self.root.brand = self.brand
#         self.root.set_pin('123456')
#         self.root.save()
#         self.root.user_permissions. \
#             add(Permission.objects.get(
#                 content_type__app_label='users', codename='has_disbursement')
#         )
#         self.root.user_permissions. \
#             add(Permission.objects.get(
#                 content_type__app_label='users', codename='accept_vodafone_onboarding')
#         )
#         self.client_user = ClientModel(client=self.root, creator=self.super_admin)
#         self.client_user.save()
#         self.checker_user = CheckerUser(
#             hierarchy=1,
#             id=15,
#             username='test_checker_user',
#             root=self.root,
#             user_type=2
#         )
#         self.checker_user.save()
#         self.level = Levels(
#             max_amount_can_be_disbursed=1200,
#             created=self.root
#         )
#         self.level.save()
#         self.checker_user.level = self.level
#         self.checker_user.save()
#         self.checker_user.user_permissions. \
#             add(Permission.objects.get(
#                 content_type__app_label='users', codename='has_disbursement')
#         )
#         self.checker_user.user_permissions. \
#             add(Permission.objects.get(
#                 content_type__app_label='users', codename='accept_vodafone_onboarding')
#         )
#         self.maker_user = MakerUser(
#             hierarchy=1,
#             id=14,
#             username='test_maker_user',
#             email='t@mk.com',
#             root=self.root,
#             user_type=1
#         )
#         self.maker_user.save()
#         self.budget = Budget(disburser=self.root, current_balance=250)
#         self.budget.save()
#         fees_setup_bank_wallet = FeeSetup(budget_related=self.budget, issuer='bc',
#                                           fee_type='f', fixed_value=20)
#         fees_setup_bank_wallet.save()
#         # fees_setup_vodafone = FeeSetup(budget_related=self.budget, issuer='vf',
#         #                                fee_type='p', percentage_value=2.25)
#         # fees_setup_vodafone.save()
#         # create auth data
#         self.checker_user.set_password('fiA#EmkjLBh9VSXy6XvFKxnR9jXt')
#         self.checker_user.save()
#         Application.objects.create(
#                 client_type=Application.CLIENT_CONFIDENTIAL, authorization_grant_type=Application.GRANT_PASSWORD,
#                 name=f"{self.checker_user.username} OAuth App", user=self.checker_user
#         )
#         # get client_secret and client_id
#         auth_data = Application.objects.get(user=self.checker_user)
#         # get auth_token
#         url = api_reverse("users:oauth2_token")
#         data = urlencode({
#             "client_id": auth_data.client_id,
#             "client_secret": auth_data.client_secret,
#             "username": self.checker_user.username,
#             "password": "fiA#EmkjLBh9VSXy6XvFKxnR9jXt",
#             "grant_type": "password"
#         })
#         response = self.client.post(url, data, content_type="application/x-www-form-urlencoded")
#         self.access_token = response.json()['access_token']
        
#         file_category = FileCategory.objects.create(
#             user_created=self.root,
#             no_of_reviews_required =1
#         )
#         self.doc = Doc.objects.create(
#             owner=self.maker_user,
#             file_category=file_category,
#             is_disbursed=False,
#             can_be_disbursed=True,
#             is_processed=True,
#         )
#         doc_review = DocReview.objects.create(
#             is_ok=True,
#             doc=self.doc,
#             user_created=self.checker_user,
#         )
#         disb_data_doc = DisbursementDocData.objects.create(
#             doc=self.doc,
#             txn_status = "200",
#             has_callback = True,
#             doc_status = "5"
#         )












  