# -*- coding: utf-8 -*-
from __future__ import unicode_literals

import csv
from django_admin_multiple_choice_list_filter.list_filters import MultipleChoiceListFilter

from django.contrib import admin
from django.contrib.admin.templatetags.admin_urls import admin_urlname
from django.contrib.contenttypes.models import ContentType
from django.shortcuts import resolve_url, render
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _
from django.http import HttpResponse
from django.db.models import Q

from disbursement.models import DisbursementData, BankTransaction
from users.models import RootUser
from .models import AmanTransaction, InstantTransaction
from .mixins import ExportCsvMixin
from core.models import AbstractBaseStatus
from utilities.date_range_filter import CustomDateRangeFilter
from django.urls import path, reverse
from django import forms
from django.contrib import messages
from django.http import HttpResponseRedirect
import datetime
from openpyxl import load_workbook
from instant_cashin.tasks import update_instant_timeouts_from_vodafone_report




class AmanTransactionTypeFilter(admin.SimpleListFilter):
    title = "Transaction Type"
    parameter_name = "transaction_type"

    def lookups(self, request, model_admin):
        return(
            ("manual_transaction", "Disbursement Data Record"),
            ("instant_transaction", "Instant Transaction")
        )

    def queryset(self, request, queryset):
        if self.value() == "manual_transaction":
            return queryset.filter(transaction_type=ContentType.objects.get_for_model(DisbursementData))
        if self.value() == 'instant_transaction':
            return queryset.filter(transaction_type=ContentType.objects.get_for_model(InstantTransaction))


class CustomStatusFilter(admin.SimpleListFilter):
    title = 'Status'
    parameter_name = 'status_choice_verbose'

    def lookups(self, request, model_admin):
        return AbstractBaseStatus.STATUS_CHOICES + [
            ('U', _("Unknown")),
        ]

    def queryset(self, request, queryset):
        value = self.value()
        if value:
            return queryset.filter(status=value)
        return queryset


class CustomRootFilter(admin.SimpleListFilter):
    title = 'Root'
    parameter_name = 'root__id'

    def lookups(self, request, model_admin):

        return RootUser.objects.all().values_list('id', 'username')

    def queryset(self, request, queryset):
        value = self.value()
        if value:
            return queryset.filter((Q(document__owner__root__id=value) | Q(from_user__root__id=value)))
        return queryset



class IssuerTypeListFilter(MultipleChoiceListFilter):
    title = 'Issuer Type'
    parameter_name = 'issuer_type__in'

    def lookups(self, request, model_admin):
        return InstantTransaction.ISSUER_TYPE_CHOICES


@admin.register(AmanTransaction)
class AmanTransactionAdmin(admin.ModelAdmin):
    """
    Admin model for Aman Instant Transaction model
    """

    list_display = ['transaction_id', 'transaction_type', 'disburser', 'bill_reference', 'is_paid']
    readonly_fields = list_display + ['original_transaction_url']
    search_fields = ['transaction_id']
    list_filter = ['is_paid', AmanTransactionTypeFilter]

    def disburser(self, instance):
        """Show the user who made the original transaction"""
        if instance.transaction_type:
            if instance.transaction_type.name == "Disbursement Data Record":
                return instance.transaction.doc.disbursed_by
            else:
                return instance.transaction.from_user

    def original_transaction_url(self, instance):
        """Create link to the original transaction"""
        if instance.transaction_type:
            if instance.transaction_type.name == "Disbursement Data Record":
                url = resolve_url(admin_urlname(DisbursementData._meta, 'change'), instance.transaction.id)
                return format_html(f"<a href='{url}'>{instance.transaction.id}</a>")
            else:
                url = resolve_url(admin_urlname(InstantTransaction._meta, 'change'), instance.transaction.uid)
                return format_html(f"<a href='{url}'>{instance.transaction.uid}</a>")

    original_transaction_url.short_description = "Go To Transaction Details"

    def has_add_permission(self, request):      # ToDo: Refactor add/delete/change to permission mixin
        return False

    def has_delete_permission(self, request, obj=None):
        if request.user.is_superuser:
            return True
        return False

    def has_change_permission(self, request, obj=None):
        return False


@admin.register(InstantTransaction)
class InstantTransactionAdmin(admin.ModelAdmin, ExportCsvMixin):
    """
    Instant Transaction Admin model for the Instant Transaction model
    """

    default_fields = [
        'uid', 'from_user', 'anon_recipient', 'status_choice_verbose', 'transaction_status_code', 'amount', 'issuer_type'
    ]
    list_display = default_fields + ['created_at', 'updated_at', 'disbursed_date', 'balance_before', 'balance_after']
    readonly_fields = default_fields + ['uid', 'created_at']
    search_fields = ['uid', 'anon_recipient']
    ordering = ['-created_at']
    list_filter = [
        ('disbursed_date', CustomDateRangeFilter),
        ('created_at', CustomDateRangeFilter),
        CustomStatusFilter,
        IssuerTypeListFilter,
        'anon_sender', 'from_user',
        CustomRootFilter,
        'is_single_step', 'transaction_status_code'
    ]
    actions = ["export_as_csv","export_bank_transactions_ids"]
    fieldsets = (
        (None, {'fields': ('from_user', )}),
        (_('Transaction Details'), {
            'fields': (
                'uid', 'reference_id', 'status_choice_verbose', 'amount', 'issuer_type', 'anon_sender', 'anon_recipient',
                'recipient_name', 'transaction_status_code', 'transaction_status_description', 'document'
            )
        }),
        (_('Important Dates'), {
            'fields': ('created_at', 'updated_at', 'disbursed_date')
        }),
        (_('Balance updates'), {
            'fields': ('balance_before', 'balance_after')
        }),
    )

    def has_module_permission(self, request):
        if request.user.is_superuser or request.user.has_perm("users.has_instant_transaction_view"):
            return True

    def has_view_permission(self, request, obj=None):
        if request.user.is_superuser or request.user.has_perm("users.has_instant_transaction_view"):
            return True

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def export_bank_transactions_ids(self, request, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'bank_transactions_ids.csv'
        writer = csv.writer(response)
        writer.writerow(["TranstactionId"])
        for instant_trx in queryset:
            bank_trx = BankTransaction.objects.filter(end_to_end=instant_trx.uid)
            if bank_trx.exists():
                transaction_id = bank_trx.first().parent_transaction.transaction_id.hex
                writer.writerow([transaction_id])

        return response

    export_bank_transactions_ids.short_description = "Export Bank Transactions Ids"

    def get_urls(self):
        urls = super().get_urls()
        new_urls = [path('upload-timeouts/', self.upload_csv),]
        return new_urls + urls

    def upload_csv(self, request):

        if request.method == "POST":
            xlsx_file = request.FILES["timeouts_upload"]
            file_name=xlsx_file.name
            date=file_name.split(".")[0]

            """
                check if the format of date is valid
            """

            try:
                datetime.datetime.strptime(date, '%Y-%m-%d')
            except ValueError:
                messages.warning(request,"Incorrect file name format, should be YYYY-MM-DD.xlsx")
                return HttpResponseRedirect(request.path_info)

            
            if not xlsx_file.name.endswith('.xlsx'):
                messages.warning(request, 'The wrong file type was uploaded')
                return HttpResponseRedirect(request.path_info)

            wb = load_workbook(xlsx_file)
            ws = wb.active
            last_row = len(list(ws.rows))
            my_dict = {}
            for row in range(2, last_row + 1):
                if int(ws["AA" + str(row)].value) < 0:
                    my_dict[ws["L" + str(row)].value] = {
                        "wallet": ws["H" + str(row)].value,
                        "success": not bool(ws["V" + str(row)].value),
                        "amount": ws["J" + str(row)].value,
                    }
            

            # print(request.POST['date_from'], request.POST['date_to'], request.user.email)
            update_instant_timeouts_from_vodafone_report.run(my_dict, date,date, request.user.email)


            # url = reverse('admin:index')
            # return HttpResponseRedirect(url)

        form = CsvImportForm()
        data = {"form": form}
        return render(request, "admin/timeouts_upload.html", data)


class CsvImportForm(forms.Form):
    timeouts_upload = forms.FileField()
    # date_from = forms.DateField()
    # date_to = forms.DateField()
